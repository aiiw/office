# -*- coding: utf-8 -*-
import openpyxl
import os
dir_path  = os.path.dirname(os.path.realpath(__file__))
test_xlsx = os.path.join(dir_path,f'''flow1.xlsx''')
print(test_xlsx)
wb = openpyxl.load_workbook(test_xlsx)

#（#“Sheet1”！A1”，“链接到Sheet1”）' ws[“A3”].style=“Hyperlink”
if '目录' not in wb.sheetnames:
	wb.create_sheet('目录')
sheet2 = wb['目录']

sheet_names = wb.sheetnames
n=0
for sheet_name in sheet_names:
	print(sheet_name)
	n=n+1
	# sheet2.cell(n,1).value = sheet_name #写单元格
	sheet2.cell(n,1).hyperlink="#'"+sheet_name+"'!A1"
	sheet2.cell(n,1).style='Hyperlink'
	# print(sheet.cell(n,1).hyperlink.target)
	wb[sheet_name].cell(1,5).hyperlink="#'目录'!A1"
	wb[sheet_name].cell(1,5).style='Hyperlink'

wb.save('flow1.xlsx')