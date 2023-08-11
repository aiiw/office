import openpyxl
work = openpyxl.load_workbook('Test.xlsx')  #打开一个工作簿
sheet_names = work.get_sheet_names() #获得sheet_name的数组
sheet1 = work.get_sheet_by_name('9.99') #以sheet名称获取sheet对象
sheet2 = work.get_sheet_by_name(sheet_names[0]) #获取名称数组后按索引，可以在不知道名称的情况下遍历
print (sheet1.cell(1,1).value) #需要注意的是单元格从1，1开始，而不是0
sheet1.cell(1,1).value = '111' #写单元格
print (sheet1.max_row) #总行数
print (sheet1.max_column)  #总列数
work.save('001T_ok.xlsx')#保存更名