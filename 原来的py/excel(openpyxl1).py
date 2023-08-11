from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Side, Border
import shutil

# 生成测试计划的excel文件
class GenerateCaseExcel(object):
    def __init__(self, file_name):
        self.file_name = file_name
        self.file_path = '/xxx/xxx/xxx/'
        self.font_title = Font(name=u"宋体", size=12, bold=True)
        self.font_body = Font(name=u"宋体", size=10)
        self.alignment_center = Alignment(horizontal='center', vertical='center', wrap_text=True)
        self.alignment_left = Alignment(horizontal='left', vertical='center', wrap_text=True)
        self.thin = Side(border_style="thin")
        self.border = Border(top=self.thin, left=self.thin, right=self.thin, bottom=self.thin)

    def generateExcel(self, basic_data, case_set_list, case_data_info):
        shutil.copy(u'/xxx/xxx/xxx/测试用例模板.xlsx', self.file_path + self.file_name + '.xlsx')

        wb = load_workbook(self.file_path + self.file_name + '.xlsx')
        # 综合评估页面
        ws_first = wb.worksheets[0]
        ws_first.cell(2, 2).value = basic_data['project_name']
        ws_first.cell(2, 4).value = basic_data['report_code']
        ws_first.cell(2, 6).value = basic_data['report_date']
        ws_first.cell(3, 2).value = basic_data['task_id']
        ws_first.cell(3, 4).value = basic_data['task_name']
        ws_first.cell(3, 6).value = basic_data['task_owner']

        ws_first.cell(4, 2).value = basic_data['task_priority']
        ws_first.cell(4, 4).value = basic_data['task_status']
        ws_first.cell(4, 6).value = basic_data['task_module']
        ws_first.cell(5, 2).value = basic_data['app_version']
        ws_first.cell(5, 4).value = basic_data['product_id']
        ws_first.cell(5, 6).value = basic_data['device_id']

        ws_first.cell(6, 2).value = basic_data['firmware_key']
        ws_first.cell(6, 4).value = basic_data['firmware_version']
        ws_first.cell(6, 6).value = basic_data['mcu_version']
        ws_first.cell(7, 2).value = basic_data['gateway_version']
        ws_first.cell(7, 4).value = basic_data['chip_module']

        ws_first.cell(8, 2).value = basic_data['task_result']
        ws_first.cell(9, 2).value = basic_data['note']
        ws_first.cell(10, 2).value = basic_data['router']
        ws_first.cell(11, 2).value = basic_data['test_mobile']

        for i in range(8, 12):
            for j in range(2, 7):
                ws_first.cell(i, j).border = self.border

        # 动态生成测试任务用例集信息
        if len(case_set_list) > 0:
            # 合并单元格处理
            merge_num = int(11) + len(case_set_list)
            ws_first.merge_cells("A12:A" + str(merge_num))
            ws_first.cell(12, 1, value="测试流程")
            ws_first.cell(12, 1).alignment = self.alignment_center
            ws_first.cell(merge_num, 1).border = self.border

            for i in range(len(case_set_list)):
                cur_row = int(12) + i
                ws_first.cell(12 + i, 2, value="用例集名称")
                ws_first.cell(12 + i, 2).alignment = self.alignment_center
                ws_first.cell(12 + i, 2).border = self.border
                ws_first.merge_cells("C" + str(cur_row) + ":D" + str(cur_row))
                ws_first.cell(12 + i, 3, value=case_set_list[i]['set_name'])
                ws_first.cell(12 + i, 3).alignment = self.alignment_center
                ws_first.cell(12 + i, 3).border = self.border
                ws_first.cell(12 + i, 4).border = self.border
                ws_first.cell(12 + i, 5, value="用例负责人")
                ws_first.cell(12 + i, 5).alignment = self.alignment_center
                ws_first.cell(12 + i, 5).border = self.border
                ws_first.cell(12 + i, 6, value=case_set_list[i]['set_owner'])
                ws_first.cell(12 + i, 6).alignment = self.alignment_center
                ws_first.cell(12 + i, 6).border = self.border

        # 测试用例集用例详细信息
        fields = "case_id,case_module,case_priority,case_tags,case_name,case_step,expect_result,case_operator,real_result,note".split(",")
        CASE_FIELD_LENGHT = 10
        CASE_FIELD_DES = ["用例编号", "模块", "优先级", "标签", "标题", "测试步骤", "期望结果", "执行人", "实际结果", "备注"]
        COLUMN_DES = {1: 'A', 2: 'B', 3: 'C', 4: 'D', 5: 'E', 6: 'F', 7: 'G', 8: 'H', 9: 'I', 10: 'J'}
        if len(case_set_list) > 0:
            for i in range(len(case_set_list)):
                # title需要是unicode类型
                ws_name = wb.create_sheet(title=case_set_list[i]['set_name'])
                # 用例第一行初始化
                for j in range(CASE_FIELD_LENGHT):
                    ws_name.cell(1, j + 1, value=CASE_FIELD_DES[j])
                    if j == 3 or j == 4 or j == 5 or j == 6 or j == 9:
                        ws_name.column_dimensions[COLUMN_DES[j + 1]].width = 35
                    else:
                        ws_name.column_dimensions[COLUMN_DES[j + 1]].width = 10
                    ws_name.cell(1, j + 1).font = self.font_title
                    ws_name.cell(1, j + 1).alignment = self.alignment_center
                    ws_name.cell(1, j + 1).border = self.border
                    ws_name.row_dimensions[1].height = 30
                if case_set_list[i]['set_name'] in case_data_info.keys() and len(case_data_info[case_set_list[i]['set_name']]) > 0:
                    self.generateTableData(ws_name, case_data_info[case_set_list[i]['set_name']], fields)

        wb.save(filename=self.file_path + self.file_name + '.xlsx')
        wb.close()

    # 生成table规则数据
    def generateTableData(self, sheet_name, data_list, fields):
        row_index = 2
        for data in data_list:
            col_index = 1
            for title in fields:
                sheet_name.cell(row=row_index, column=col_index, value=data[title])
                sheet_name.cell(row=row_index, column=col_index).border = self.border
                sheet_name.cell(row=row_index, column=col_index).font = self.font_body
                sheet_name.row_dimensions[row_index].height = 25
                if col_index == 5 or col_index == 6 or col_index == 7 or col_index == 10:
                    sheet_name.cell(row=row_index, column=col_index).alignment = self.alignment_left
                else:
                    sheet_name.cell(row=row_index, column=col_index).alignment = self.alignment_center
                col_index += 1
            row_index += 1