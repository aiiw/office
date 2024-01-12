import xlwings as xw

#连接到excel
workbook = xw.Book('你的excel文件的路径')#连接excel文件
#连接到指定单元格
data_range = workbook.sheets('Sheet1').range('A1')
#写入数据
data_range.value = ['a','b','c']
#保存
workbook.save()

#########################################################################################################
import xlrd #读取数据
import xlwt #写入数据
import xlutils.copy #操作excel
# 通过xlrd读取数据

#打开excel文件
workbook = xlrd.open_workbook('你的excel文件的路径')
#获取表单
worksheet = workbook.sheet_by_index(0)
#读取数据
data = worksheet.cell_value(0,0)

# 通过xlwt写入数据
#新建excel
wb = xlwt.Workbook()
#添加工作薄
sh = wb.add_sheet('Sheet1')
#写入数据
sh.write(0,0,'abc')
#保存文件
wb.save('myexcel.xls')


#打开excel文件
book = xlrd.open_workbook('你的excel文件的路径')
#复制一份
new_book = copy(book)
#拿到工作薄
worksheet = new_book.getsheet(0)
#写入数据
worksheet.write(0,0,'mydata')
#保存
new_book.save()
##############################################################################

import openpyxl
 # 新建文件
 workbook = openpyxl.Workbook() 
 # 写入文件
 sheet = workbook.activesheet['A1']='A1'
 # 保存文件 
 workbook.save('你的excel保存路径')



#打开存在
dir_path  = os.path.dirname(os.path.realpath(__file__))
test_xlsx = os.path.join(dir_path,f'''flow1.xlsx''')
wb = openpyxl.load_workbook(test_xlsx)

'''
当使用 openpyxl 库处理 Excel 文件时，以下是一些常用的方法：

打开 Excel 文件：

load_workbook(filename): 打开指定文件名的 Excel 文件并返回 Workbook 对象。
获取工作表：

workbook.sheetnames: 返回工作簿中所有工作表的名称列表。
workbook[sheetname]: 返回指定名称的工作表对象。
读取单元格数据：

worksheet.cell(row, column): 返回指行和列的单元格对象。
cell.value: 返回单元格的值。
写入单元格数据：

worksheet.cell(row, column, value): 将指定值写入指定和列的单元格。
获取行和列的围：

worksheet.max_row: 返回工作表中最大行数。
worksheet.max_column: 返回工作表中最大列数。
迭代行或列：

worksheet.iter_rows(min_row, max_row, min_col, max_col): 返回指定行和列范围内的单格迭代器。
worksheet.iter_cols(min_row, max_row, min_col, max_col): 返回指定行和列范围的单元格迭代器。
保存 Excel 文件：

workbook.save(filename): 将对 Excel 文件的修改保存到定的文件名。
关 Excel 文件：

workbook.close(): 关闭已打开的 Excel 文件。
这些是 openpyxl 库中的一些常用方法，可以帮助你读取、写入和处理 Excel 文件。你可以根据具体需求使用这些方法来操作 Excel 数据。'''