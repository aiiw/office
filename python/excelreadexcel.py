# import xlrd
# book=xlrd.open_workbook('getimaat.xlsx') #打开Excel
# sheet=book.sheet_by_name('max') #根据编号获取sheet页
# #sheet=book.sheet_by_name('sheet1') #也可以根据sheet页名字获取sheet页
# #、第一步将excel表读取到二维表---->list2

# nrows=sheet.nrows
# list1=[]
# list2=[]
# list3=[]
# list4=[]
# for i in range(1,nrows,1):  # 将excel二维表－－〈放到list2 二维表〉
#    list1=sheet.row_values(i)
#    list2.append(list1)
# #print(list2)
# m=0  #m 读取每一行的一个字段
# #x=0
# x2=0 #将分群+0000
# #这种方式是传统的for
# for j in list2: #j为一行 视同 tabel中的row j[0] 视同字段
	
# 	#m=0
# 	m=int(j[1])
# 	x=1
# 	list3=[]

# 	for p in range(m): #p为记数器

# 		x=x+1
# 		x1=j[0]
# 		x2=j[3]+x
# 		#str1=('%04d'%(x2)) #这个是传统的写法，新的为format
# 		str1='{:0>4d}'.format(int(x2)) #测试 了下这里的参数与模板的数据类型需要一样
# 		list3=[j[0],j[0]+str1]
# 		list4.append(list3)
# print(list4)
# #3、将处理 好的数据写入到excel
# list4=[[1,2,3],[4,5,6]] 典型的二维表，如果写入xw要求每个列表的的长度一致
# import xlwings as xw
# app=xw.App(visible=False,add_book=False)
# book=app.books.add()
# sheet=book.sheets['sheet1']
# sheet.range('a1').value=[[1,2],[1,2],[123,33]]
# book.save('77.xlsx')
# book.close()
# app.quit()
		
import openpyxl as app
from openpyxl import load_workbook

# workbook = load_workbook(filename = "77.xlsx")
# sheet1=workbook.active
# cell0=sheet1["A"]
# print(type(cell0),cell0)
# # 其实以元组记录cell对象,如果是列或行则一维,如果区域则二维aaa
# # for row in cell0:
# # 	for col in row:
# # 		print(col.value)
# cell1=sheet1.cell(2,3)
# cell1.value="aiiw"
# sheet1.cell(1,2).value="aaiwrewre"
# sheet1.append(['2','3','4'])
# workbook.save(filename = "77.xlsx")



wb = app.Workbook()
sheet1=wb.active
sheet1["A1"]="aiiw"
sheet1.auto_filter.ref=sheet1["A1"]
wb.save("pyxl1.xlsx")
# print(sheet.cell(0,0).value) #获取到指定单元格的内容
# print(sheet.cell(0,1).value) #获取到指定单元格的内容

# print(sheet.row_values(0))  #获取到整行的内容
# print(sheet.col_values(0))   #获取到整列的内容

# for i in range(sheet.nrows):  #循环获取每行的内容
#     print(sheet.row_values(i))
